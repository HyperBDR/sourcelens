import json
import mimetypes
import time
import zipfile
from io import BytesIO
from pathlib import Path

from .conversion_queue import ConversionJob
from .conversion_queue import conversion_queue_from_context
from .datasource_manifest import manifest_local_path
from .datasource_manifest import should_skip_dir
from .path_rules import is_excluded_path
from .path_rules import is_sidecar_dir
from .path_rules import safe_filename
from .path_rules import sidecar_path
from .path_rules import source_sha256

DOCUMENT_EXTENSIONS = {".pdf", ".docx", ".pptx", ".xlsx"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
EMBEDDED_IMAGE_PREFIXES = {
    ".docx": "word/media/",
    ".pptx": "ppt/media/",
    ".xlsx": "xl/media/",
}
PROMPT_VERSION = "image-search-v1"
DEFAULT_MAX_IMAGES = 100
DEFAULT_MAX_FILE_SIZE_MB = 100
DEFAULT_MAX_PAGES = 500
DEFAULT_IMAGE_JPEG_QUALITY = 82
DEFAULT_IMAGE_MAX_DIMENSION = 1600
DEFAULT_MIN_IMAGE_BYTES = 5 * 1024
DEFAULT_MIN_IMAGE_DIMENSION = 64
DEFAULT_PDF_MAX_IMAGES_PER_PAGE = 3
DEFAULT_PDF_MAX_PAGES = 30
DEFAULT_PDF_MIN_TEXT_CHARS = 30
DEFAULT_PDF_MIN_IMAGE_AREA_RATIO = 0.08
DEFAULT_PDF_RENDER_DPI = 144
IMAGE_BLANK_VARIANCE_THRESHOLD = 8.0
DEFAULT_TOKEN_CHARS = 4
DETAIL_ITEMS_LIMIT = 200


class ConversionOutput:
    """One converter output."""

    def __init__(
        self,
        text="",
        stats=None,
        cost=None,
        skipped=False,
        reason="",
    ):
        self.text = text
        self.stats = stats or {}
        self.cost = cost or {}
        self.skipped = skipped
        self.reason = reason


class BaseConverter:
    """Base file converter."""

    name = "base"
    extensions = set()

    def convert(self, path, context):
        """Convert one file into searchable text."""

        raise NotImplementedError

    def version(self):
        """Return converter version metadata."""

        return {"name": self.name, "version": ""}


class MarkItDownDocumentConverter(BaseConverter):
    """Document converter backed by MarkItDown."""

    name = "markitdown"
    extensions = DOCUMENT_EXTENSIONS

    def convert(self, path, context):
        """Convert a document file into Markdown text."""

        try:
            from markitdown import MarkItDown
        except Exception as exc:
            raise RuntimeError("MARKITDOWN_NOT_AVAILABLE") from exc

        result = MarkItDown().convert(str(path))
        text = getattr(result, "text_content", "") or str(result)
        stats = document_stats(path)
        cost = empty_cost_stats()
        image_context = document_image_context(context, path, text)
        embedded = convert_embedded_images(path, image_context)
        if embedded["markdown"]:
            text = f"{text.rstrip()}\n\n{embedded['markdown']}"
        stats.update(embedded["stats"])
        merge_cost_stats(cost, embedded["cost"])
        return ConversionOutput(text=text, stats=stats, cost=cost)

    def version(self):
        """Return MarkItDown version metadata."""

        try:
            import markitdown

            return {"name": self.name, "version": markitdown.__version__}
        except Exception:
            return {"name": self.name, "version": "unknown"}


class GatewayImageConverter(BaseConverter):
    """Image converter backed by the LensNode AI gateway."""

    name = "gateway_vision"
    extensions = IMAGE_EXTENSIONS

    def convert(self, path, context):
        """Convert an image file into a searchable description."""

        context = standalone_image_context(context, path)
        prepared = prepare_image_for_model(path, context)
        if prepared.get("skipped"):
            return ConversionOutput(
                stats=prepared.get("stats") or {},
                skipped=True,
                reason=prepared.get("reason") or "",
            )

        content, usage = describe_image_bytes(
            prepared["bytes"],
            prepared["mime_type"],
            context,
        )
        stats = {
            **(prepared.get("stats") or {}),
            "images_total": 1,
            "images_recognized": 1 if content else 0,
        }
        return ConversionOutput(
            text=content,
            stats=stats,
            cost=model_cost_stats(usage, content),
        )

    def version(self):
        """Return gateway converter version metadata."""

        return {"name": self.name, "version": PROMPT_VERSION}


class ConverterRegistry:
    """Registry for pluggable converters."""

    def __init__(self):
        self._converters = []

    def register(self, converter):
        """Register one converter."""

        self._converters.append(converter)

    def converter_for(self, path):
        """Return a converter for a file path."""

        suffix = Path(path).suffix.lower()
        for converter in self._converters:
            if suffix in converter.extensions:
                return converter
        return None


CONVERTERS = ConverterRegistry()
CONVERTERS.register(MarkItDownDocumentConverter())
CONVERTERS.register(GatewayImageConverter())


def post_process_documents(context, sync_result, emit=None):
    """Convert changed datasource files into searchable sidecars."""

    conversion = context.get("conversion") or {}
    if not conversion_enabled(conversion):
        return {
            "candidates": 0,
            "converted": 0,
            "success": 0,
            "skipped": 0,
            "failed": 0,
            "markdown": 0,
            "deleted_sidecars": 0,
            "chars": 0,
            "estimated_tokens": 0,
            "images_recognized": 0,
            "images_skipped": 0,
            "images_blank": 0,
            "images_duplicate": 0,
            "images_compressed": 0,
            "embedded_images_total": 0,
            "embedded_images_recognized": 0,
            "embedded_images_skipped": 0,
            "embedded_images_duplicate": 0,
            "embedded_images_blank": 0,
            "pdf_pages": 0,
            "pdf_pages_processed": 0,
            "pdf_pages_with_text": 0,
            "pdf_scanned_pages": 0,
            "pdf_images_total": 0,
            "pdf_images_recognized": 0,
            "pdf_images_skipped": 0,
            "pdf_rendered_pages": 0,
            "xlsx_files": 0,
            "sheets": 0,
            "rows": 0,
            "truncated_files": 0,
            "cost": empty_cost_stats(),
            "warnings": [],
            "items": [],
            "items_truncated": 0,
            "details": {},
            "details_truncated": {},
        }

    target = Path(context["target_path"]).resolve()
    excluded_roots = context.get("excluded_datasource_roots") or []
    warnings = []
    candidates = conversion_candidates(
        target,
        sync_result.items,
        context.get("datasource_uuid") or "",
        excluded_roots,
        conversion,
    )
    total = len(candidates)
    summary = {
        "candidates": total,
        "converted": 0,
        "success": 0,
        "skipped": 0,
        "failed": 0,
        "markdown": 0,
        "deleted_sidecars": 0,
        "chars": 0,
        "estimated_tokens": 0,
        "images_recognized": 0,
        "images_skipped": 0,
        "images_blank": 0,
        "images_duplicate": 0,
        "images_compressed": 0,
        "embedded_images_total": 0,
        "embedded_images_recognized": 0,
        "embedded_images_skipped": 0,
        "embedded_images_duplicate": 0,
        "embedded_images_blank": 0,
        "pdf_pages": 0,
        "pdf_pages_processed": 0,
        "pdf_pages_with_text": 0,
        "pdf_scanned_pages": 0,
        "pdf_images_total": 0,
        "pdf_images_recognized": 0,
        "pdf_images_skipped": 0,
        "pdf_rendered_pages": 0,
        "xlsx_files": 0,
        "sheets": 0,
        "rows": 0,
        "truncated_files": 0,
        "cost": empty_cost_stats(),
        "warnings": warnings,
        "items": [],
        "items_truncated": 0,
        "details": {},
        "details_truncated": {},
    }
    emit_conversion(
        emit,
        "conversion_plan",
        "running",
        f"Prepared {total} files for conversion.",
        summary,
        total=total,
        current=0,
    )

    max_images = int(conversion.get("max_images") or DEFAULT_MAX_IMAGES)
    image_count = 0
    image_digests = set()
    jobs = []
    for index, item in enumerate(candidates, start=1):
        path = target / manifest_local_path(item)
        if is_image_path(path):
            image_count += 1
            if image_count > max_images:
                stats = image_skip_stats("CONVERSION_MAX_IMAGES_EXCEEDED")
                summary["skipped"] += 1
                summary["images_skipped"] += 1
                warnings.append("CONVERSION_MAX_IMAGES_EXCEEDED")
                append_conversion_detail(
                    summary,
                    item,
                    "skipped",
                    "CONVERSION_MAX_IMAGES_EXCEEDED",
                    stats,
                )
                emit_conversion(
                    emit,
                    "conversion_progress",
                    "running",
                    f"Converted {index}/{total} datasource files.",
                    summary,
                    total=total,
                    current=index,
                    current_file=manifest_local_path(item),
                    current_status="skipped",
                    current_reason="CONVERSION_MAX_IMAGES_EXCEEDED",
                    current_stats=stats,
                )
                continue
            digest = source_sha256(path)
            if digest in image_digests:
                stats = image_skip_stats("IMAGE_DUPLICATE")
                summary["skipped"] += 1
                summary["images_skipped"] += 1
                summary["images_duplicate"] += 1
                append_conversion_detail(
                    summary,
                    item,
                    "skipped",
                    "IMAGE_DUPLICATE",
                    stats,
                )
                emit_conversion(
                    emit,
                    "conversion_progress",
                    "running",
                    f"Converted {index}/{total} datasource files.",
                    summary,
                    total=total,
                    current=index,
                    current_file=manifest_local_path(item),
                    current_status="skipped",
                    current_reason="IMAGE_DUPLICATE",
                    current_stats=stats,
                )
                continue
            image_digests.add(digest)
        jobs.append(
            ConversionJob(
                index=index,
                total=total,
                item=item,
                path=path,
            )
        )
    queue = conversion_queue_from_context(context)

    def handle_job(job):
        return convert_job(job, target, context)

    for job, output in queue.run(jobs, handle_job):
        path = job.path
        item = job.item
        index = job.index
        current_status = "converted"
        current_reason = ""
        current_stats = {}
        try:
            if isinstance(output, Exception):
                raise output
            result = output
            if result.get("skipped"):
                current_status = "skipped"
                current_reason = (
                    result.get("reason") or result.get("warning") or ""
                )
                current_stats = result.get("stats") or {}
                summary["skipped"] += 1
                if result.get("warning"):
                    warnings.append(result["warning"])
                append_conversion_detail(
                    summary,
                    item,
                    "skipped",
                    current_reason,
                    current_stats,
                )
                merge_summary_stats(summary, current_stats)
            else:
                current_stats = {
                    **(result.get("stats") or {}),
                    "cost": result.get("cost") or {},
                }
                summary["converted"] += 1
                summary["success"] += 1
                summary["markdown"] += 1
                append_conversion_detail(
                    summary,
                    item,
                    "converted",
                    "",
                    current_stats,
                )
                summary["chars"] += int(result.get("chars") or 0)
                summary["estimated_tokens"] += int(
                    result.get("estimated_tokens") or 0
                )
                summary["images_recognized"] += int(
                    result.get("images_recognized") or 0
                )
                merge_summary_stats(summary, result.get("stats") or {})
                merge_cost_stats(summary["cost"], result.get("cost") or {})
        except Exception as exc:
            current_status = "failed"
            current_reason = str(exc)
            summary["failed"] += 1
            warnings.append("CONVERSION_FILE_FAILED")
            write_failed_meta(target, path, item, context, str(exc))
            append_conversion_detail(summary, item, "failed", str(exc))
        emit_conversion(
            emit,
            "conversion_progress",
            "running",
            f"Converted {index}/{total} datasource files.",
            summary,
            total=total,
            current=index,
            current_file=manifest_local_path(item),
            current_status=current_status,
            current_reason=current_reason,
            current_stats=current_stats,
        )

    summary["details"] = conversion_details_by_metric(summary["items"])
    summary["details_truncated"] = conversion_details_truncated(summary)
    summary["warnings"] = list(dict.fromkeys(warnings))
    emit_conversion(
        emit,
        "conversion_manifest",
        "done",
        "Datasource conversion completed.",
        summary,
        total=total,
        current=total,
    )
    return summary


def append_conversion_detail(summary, item, status, reason="", stats=None):
    """Append a compact conversion item detail."""

    if len(summary["items"]) >= DETAIL_ITEMS_LIMIT:
        summary["items_truncated"] += 1
        return
    local_path = manifest_local_path(item)
    stats = dict(stats or {})
    summary["items"].append(
        {
            "status": status,
            "path": local_path,
            "name": item.get("name") or Path(local_path).name,
            "extension": Path(local_path).suffix.lower().lstrip("."),
            "reason": reason,
            "stats": stats,
        }
    )


def conversion_details_by_metric(items):
    """Return conversion details grouped by summary metric."""

    details = {
        "candidates": [],
        "converted": [],
        "success": [],
        "failed": [],
        "markdown": [],
        "skipped": [],
        "xlsx_files": [],
        "sheets": [],
        "rows": [],
        "model_calls": [],
        "estimated_tokens": [],
        "total_tokens": [],
        "images_skipped": [],
        "images_recognized": [],
        "images_blank": [],
        "images_duplicate": [],
        "images_compressed": [],
        "embedded_images_total": [],
        "embedded_images_recognized": [],
        "embedded_images_skipped": [],
        "embedded_images_duplicate": [],
        "embedded_images_blank": [],
        "pdf_pages": [],
        "pdf_pages_processed": [],
        "pdf_pages_with_text": [],
        "pdf_scanned_pages": [],
        "pdf_images_total": [],
        "pdf_images_recognized": [],
        "pdf_images_skipped": [],
        "pdf_rendered_pages": [],
    }
    truncated = {key: 0 for key in details}
    for item in items or []:
        status = item.get("status")
        _append_conversion_metric_detail(
            details,
            truncated,
            "candidates",
            item,
        )
        if status == "converted":
            for key in ["converted", "success", "markdown"]:
                _append_conversion_metric_detail(details, truncated, key, item)
        elif status == "skipped":
            _append_conversion_metric_detail(
                details,
                truncated,
                "skipped",
                item,
            )
        elif status == "failed":
            _append_conversion_metric_detail(
                details,
                truncated,
                "failed",
                item,
            )
        if item.get("extension") == "xlsx":
            for key in ["xlsx_files", "sheets", "rows"]:
                _append_conversion_metric_detail(details, truncated, key, item)
        cost = (item.get("stats") or {}).get("cost") or {}
        if int(cost.get("model_calls") or 0) > 0:
            _append_conversion_metric_detail(
                details,
                truncated,
                "model_calls",
                item,
            )
        if int(cost.get("estimated_tokens") or 0) > 0:
            _append_conversion_metric_detail(
                details,
                truncated,
                "estimated_tokens",
                item,
            )
        if int(cost.get("total_tokens") or 0) > 0:
            _append_conversion_metric_detail(
                details,
                truncated,
                "total_tokens",
                item,
            )
        stats = item.get("stats") or {}
        for key in [
            "images_skipped",
            "images_recognized",
            "images_blank",
            "images_duplicate",
            "images_compressed",
            "embedded_images_total",
            "embedded_images_recognized",
            "embedded_images_skipped",
            "embedded_images_duplicate",
            "embedded_images_blank",
            "pdf_pages",
            "pdf_pages_processed",
            "pdf_pages_with_text",
            "pdf_scanned_pages",
            "pdf_images_total",
            "pdf_images_recognized",
            "pdf_images_skipped",
            "pdf_rendered_pages",
        ]:
            if int(stats.get(key) or 0) > 0:
                _append_conversion_metric_detail(
                    details,
                    truncated,
                    key,
                    item,
                )
    return {key: value for key, value in details.items() if value}


def conversion_details_truncated(summary):
    """Return conversion detail truncation counts grouped by metric."""

    truncated = {}
    if summary.get("items_truncated"):
        truncated["candidates"] = int(summary["items_truncated"])
    return truncated


def _append_conversion_metric_detail(details, truncated, key, item):
    """Append one conversion detail to a metric group."""

    if len(details[key]) >= DETAIL_ITEMS_LIMIT:
        truncated[key] += 1
        return
    details[key].append(item)


def conversion_enabled(conversion):
    """Return whether any conversion option is enabled."""

    return bool(
        conversion.get("document")
        or conversion.get("image")
        or conversion.get("embedded_image")
    )


def conversion_candidates(
    target,
    items,
    datasource_uuid,
    excluded_roots,
    conversion,
):
    """Return manifest items eligible for conversion."""

    candidates = []
    for item in items or []:
        if item.get("status") == "deleted":
            continue
        local_path = manifest_local_path(item)
        if not local_path:
            continue
        path = (target / local_path).resolve()
        if not path.is_file() or is_excluded_path(path, excluded_roots):
            continue
        if not is_convertible(path, conversion):
            continue
        if has_foreign_marker(
            path.parent,
            target,
            datasource_uuid,
            excluded_roots,
        ):
            continue
        candidates.append(item)
    return candidates


def has_foreign_marker(parent, target, datasource_uuid, excluded_roots):
    """Return whether path is under a foreign datasource root."""

    for directory in [parent, *parent.parents]:
        if not is_under_target(directory, target):
            break
        if should_skip_dir(directory, datasource_uuid, excluded_roots):
            return True
    return False


def is_under_target(path, target):
    """Return whether path is within the current datasource target."""

    try:
        Path(path).resolve().relative_to(Path(target).resolve())
        return True
    except ValueError:
        return False


def is_convertible(path, conversion):
    """Return whether path is enabled for conversion."""

    suffix = Path(path).suffix.lower()
    if conversion.get("document") and suffix in DOCUMENT_EXTENSIONS:
        return True
    if conversion.get("image") and suffix in IMAGE_EXTENSIONS:
        return True
    return False


def convert_job(job, target, context):
    """Convert one queued job."""

    return convert_one(target, job.path, job.item, context)


def convert_one(target, path, item, context):
    """Convert one file if fingerprint requires it."""

    limits = conversion_limits(context.get("conversion") or {})
    size = path.stat().st_size
    if size > limits["max_file_size"]:
        write_skipped_meta(target, path, item, context, "FILE_TOO_LARGE")
        return {"skipped": True, "reason": "FILE_TOO_LARGE"}

    digest = source_sha256(path)
    fingerprint = conversion_fingerprint(path, digest, context)
    meta_path = sidecar_path(path) / "meta.json"
    content_path = sidecar_path(path) / "content.md"
    meta = read_json(meta_path)
    if (
        meta.get("conversion", {}).get("fingerprint") == fingerprint
        and meta.get("conversion", {}).get("status") == "success"
        and content_path.is_file()
    ):
        return {"skipped": True, "reason": "UNCHANGED"}

    converter = CONVERTERS.converter_for(path)
    if converter is None:
        write_skipped_meta(target, path, item, context, "UNSUPPORTED_TYPE")
        return {"skipped": True, "reason": "UNSUPPORTED_TYPE"}

    if is_image_path(path):
        if not vision_configured(context):
            write_skipped_meta(
                target,
                path,
                item,
                context,
                "VISION_NOT_CONFIGURED",
            )
            return {
                "skipped": True,
                "reason": "VISION_NOT_CONFIGURED",
                "warning": "VISION_NOT_CONFIGURED",
            }

    output = converter.convert(path, context)
    if output.skipped:
        write_skipped_meta(target, path, item, context, output.reason)
        return {
            "skipped": True,
            "reason": output.reason,
            "stats": output.stats,
        }

    text = output.text
    stats = output.stats
    cost = output.cost
    images_recognized = int(stats.get("images_recognized") or 0)

    sidecar_path(path).mkdir(parents=True, exist_ok=True)
    content = content_markdown(target, path, context, text)
    content_path.write_text(content, encoding="utf-8")
    write_success_meta(
        target,
        path,
        item,
        context,
        digest,
        fingerprint,
        len(text),
        images_recognized,
        stats,
    )
    return {
        "chars": len(text),
        "estimated_tokens": estimate_tokens(text),
        "images_recognized": images_recognized,
        "stats": stats,
        "cost": cost,
    }


def conversion_limits(conversion):
    """Return conversion limits in bytes/counts."""

    size_mb = int(
        conversion.get("max_file_size_mb") or DEFAULT_MAX_FILE_SIZE_MB
    )
    return {
        "max_file_size": size_mb * 1024 * 1024,
        "max_pages": int(conversion.get("max_pages") or DEFAULT_MAX_PAGES),
    }


def document_stats(path):
    """Return document-specific conversion stats."""

    suffix = Path(path).suffix.lower()
    if suffix == ".xlsx":
        return xlsx_stats(path)
    return {"pages": 0}


def xlsx_stats(path):
    """Return XLSX workbook stats when openpyxl is available."""

    try:
        from openpyxl import load_workbook
    except Exception:
        return {
            "xlsx_files": 1,
            "sheets": 0,
            "rows": 0,
            "columns": 0,
            "truncated": False,
            "warning": "OPENPYXL_NOT_AVAILABLE",
        }

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    total_rows = 0
    max_columns = 0
    try:
        for sheet in workbook.worksheets:
            rows = int(sheet.max_row or 0)
            columns = int(sheet.max_column or 0)
            total_rows += rows
            max_columns = max(max_columns, columns)
            sheets.append(
                {
                    "name": sheet.title,
                    "rows": rows,
                    "columns": columns,
                }
            )
    finally:
        workbook.close()
    return {
        "xlsx_files": 1,
        "sheets": len(sheets),
        "rows": total_rows,
        "columns": max_columns,
        "sheet_stats": sheets,
        "truncated": False,
    }


def estimate_tokens(text):
    """Return a rough token estimate for conversion cost summaries."""

    return max(0, int(len(text or "") / DEFAULT_TOKEN_CHARS))


def empty_cost_stats():
    """Return empty conversion cost stats."""

    return {
        "estimated_tokens": 0,
        "prompt_tokens": 0,
        "completion_tokens": 0,
        "total_tokens": 0,
        "model_calls": 0,
    }


def model_cost_stats(usage, text):
    """Return model usage stats for one conversion result."""

    usage = usage or {}
    total_tokens = int(usage.get("total_tokens") or 0)
    if total_tokens <= 0:
        total_tokens = estimate_tokens(text)
    return {
        "estimated_tokens": estimate_tokens(text),
        "prompt_tokens": int(usage.get("prompt_tokens") or 0),
        "completion_tokens": int(usage.get("completion_tokens") or 0),
        "total_tokens": total_tokens,
        "model_calls": 1,
    }


def merge_cost_stats(target, source):
    """Merge model cost stats into a summary bucket."""

    for key in empty_cost_stats():
        target[key] = int(target.get(key) or 0) + int(source.get(key) or 0)


def merge_summary_stats(summary, stats):
    """Merge converter stats into conversion summary."""

    if not stats:
        return
    summary["xlsx_files"] += int(stats.get("xlsx_files") or 0)
    summary["sheets"] += int(stats.get("sheets") or 0)
    summary["rows"] += int(stats.get("rows") or 0)
    summary["images_skipped"] += int(stats.get("images_skipped") or 0)
    summary["images_blank"] += int(stats.get("images_blank") or 0)
    summary["images_duplicate"] += int(stats.get("images_duplicate") or 0)
    summary["images_compressed"] += int(stats.get("images_compressed") or 0)
    summary["embedded_images_total"] += int(
        stats.get("embedded_images_total") or 0
    )
    summary["embedded_images_recognized"] += int(
        stats.get("embedded_images_recognized") or 0
    )
    summary["embedded_images_skipped"] += int(
        stats.get("embedded_images_skipped") or 0
    )
    summary["embedded_images_duplicate"] += int(
        stats.get("embedded_images_duplicate") or 0
    )
    summary["embedded_images_blank"] += int(
        stats.get("embedded_images_blank") or 0
    )
    for key in [
        "pdf_pages",
        "pdf_pages_processed",
        "pdf_pages_with_text",
        "pdf_scanned_pages",
        "pdf_images_total",
        "pdf_images_recognized",
        "pdf_images_skipped",
        "pdf_rendered_pages",
    ]:
        summary[key] += int(stats.get(key) or 0)
    if stats.get("truncated"):
        summary["truncated_files"] += 1


def conversion_fingerprint(path, digest, context):
    """Return conversion fingerprint for one source file."""

    conversion = context.get("conversion") or {}
    payload = {
        "source_sha256": digest,
        "options": {
            "document": bool(conversion.get("document")),
            "image": bool(conversion.get("image")),
            "embedded_image": bool(conversion.get("embedded_image")),
            "document_model_ref": conversion.get("document_model_ref") or "",
            "image_jpeg_quality": int(
                conversion.get("image_jpeg_quality")
                or DEFAULT_IMAGE_JPEG_QUALITY
            ),
            "image_max_dimension": int(
                conversion.get("image_max_dimension")
                or DEFAULT_IMAGE_MAX_DIMENSION
            ),
            "min_image_bytes": int(
                conversion.get("min_image_bytes") or DEFAULT_MIN_IMAGE_BYTES
            ),
            "min_image_dimension": int(
                conversion.get("min_image_dimension")
                or DEFAULT_MIN_IMAGE_DIMENSION
            ),
            "pdf_extract_images": conversion.get("pdf_extract_images")
            is not False,
            "pdf_extract_images_on_text_pages": bool(
                conversion.get("pdf_extract_images_on_text_pages")
            ),
            "pdf_max_images_per_page": int(
                conversion.get("pdf_max_images_per_page")
                or DEFAULT_PDF_MAX_IMAGES_PER_PAGE
            ),
            "pdf_max_pages": int(
                conversion.get("pdf_max_pages") or DEFAULT_PDF_MAX_PAGES
            ),
            "pdf_min_text_chars": int(
                conversion.get("pdf_min_text_chars")
                or DEFAULT_PDF_MIN_TEXT_CHARS
            ),
            "pdf_min_image_area_ratio": float(
                conversion.get("pdf_min_image_area_ratio")
                or DEFAULT_PDF_MIN_IMAGE_AREA_RATIO
            ),
            "pdf_render_dpi": int(
                conversion.get("pdf_render_dpi") or DEFAULT_PDF_RENDER_DPI
            ),
            "pdf_render_scanned_pages": bool(
                conversion.get("pdf_render_scanned_pages")
            ),
        },
        "tool": converter_version(path),
        "model_ref": conversion.get("vision_model_ref") or "",
        "prompt_version": PROMPT_VERSION,
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    import hashlib

    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def converter_version(path):
    """Return converter version metadata."""

    converter = CONVERTERS.converter_for(path)
    if converter is not None:
        return converter.version()
    return {"name": "none", "version": ""}


def convert_embedded_images(path, context):
    """Recognize embedded document images."""

    conversion = context.get("conversion") or {}
    if not conversion.get("embedded_image") or not is_document_path(path):
        return {"markdown": "", "stats": {}, "cost": empty_cost_stats()}
    if Path(path).suffix.lower() == ".pdf":
        return convert_pdf_images(path, context)

    prefix = EMBEDDED_IMAGE_PREFIXES.get(Path(path).suffix.lower())
    if not prefix:
        return {"markdown": "", "stats": {}, "cost": empty_cost_stats()}

    stats = {
        "embedded_images_total": 0,
        "embedded_images_recognized": 0,
        "embedded_images_skipped": 0,
        "embedded_images_duplicate": 0,
        "embedded_images_blank": 0,
    }
    cost = empty_cost_stats()
    descriptions = []
    if not vision_configured(context):
        stats["embedded_images_skipped"] = count_embedded_image_entries(
            path,
            prefix,
        )
        return {"markdown": "", "stats": stats, "cost": cost}

    seen = set()
    max_images = int(conversion.get("max_images") or DEFAULT_MAX_IMAGES)
    assets_dir = sidecar_path(path) / "assets"
    try:
        with zipfile.ZipFile(path) as archive:
            entries = embedded_image_entries(archive, prefix)
            for index, name in enumerate(entries, start=1):
                stats["embedded_images_total"] += 1
                if index > max_images:
                    stats["embedded_images_skipped"] += 1
                    stats["images_skipped"] = int(
                        stats.get("images_skipped") or 0
                    ) + 1
                    continue
                raw = archive.read(name)
                digest = source_bytes_sha256(raw)
                if digest in seen:
                    stats["embedded_images_skipped"] += 1
                    stats["embedded_images_duplicate"] += 1
                    stats["images_skipped"] = int(
                        stats.get("images_skipped") or 0
                    ) + 1
                    stats["images_duplicate"] = int(
                        stats.get("images_duplicate") or 0
                    ) + 1
                    continue
                seen.add(digest)
                result = convert_one_embedded_image(
                    path,
                    assets_dir,
                    name,
                    raw,
                    context,
                )
                merge_embedded_image_stats(stats, result["stats"])
                merge_cost_stats(cost, result["cost"])
                if result["description"]:
                    descriptions.append(result)
    except (OSError, zipfile.BadZipFile):
        return {"markdown": "", "stats": stats, "cost": cost}

    return {
        "markdown": embedded_images_markdown(descriptions),
        "stats": stats,
        "cost": cost,
    }


def embedded_image_entries(archive, prefix):
    """Return supported embedded image archive entries."""

    entries = []
    for name in archive.namelist():
        lower = name.lower()
        if not lower.startswith(prefix):
            continue
        if Path(lower).suffix not in IMAGE_EXTENSIONS:
            continue
        entries.append(name)
    return sorted(entries)


def count_embedded_image_entries(path, prefix):
    """Return the number of supported embedded image entries."""

    try:
        with zipfile.ZipFile(path) as archive:
            return len(embedded_image_entries(archive, prefix))
    except (OSError, zipfile.BadZipFile):
        return 0


def convert_one_embedded_image(source_path, assets_dir, name, raw, context):
    """Recognize one embedded image and return markdown data."""

    assets_dir.mkdir(parents=True, exist_ok=True)
    suffix = Path(name).suffix.lower() or ".png"
    digest = source_bytes_sha256(raw)
    asset_path = assets_dir / f"embedded_{digest[:12]}{suffix}"
    asset_path.write_bytes(raw)
    prepared = prepare_image_for_model(asset_path, context)
    if prepared.get("skipped"):
        return {
            "description": "",
            "stats": embedded_image_skip_stats(prepared),
            "cost": empty_cost_stats(),
        }

    content, usage = describe_image_bytes(
        prepared["bytes"],
        prepared["mime_type"],
        context,
    )
    stats = {
        **(prepared.get("stats") or {}),
        "embedded_images_recognized": 1 if content else 0,
        "images_recognized": 1 if content else 0,
    }
    return {
        "source": name,
        "asset": asset_path.name,
        "description": content,
        "stats": stats,
        "cost": model_cost_stats(usage, content),
    }


def embedded_image_skip_stats(prepared):
    """Return embedded image skip stats from a preprocessing result."""

    reason = prepared.get("reason") or ""
    source_stats = prepared.get("stats") or {}
    stats = {
        **source_stats,
        "embedded_images_skipped": 1,
    }
    if reason == "IMAGE_BLANK":
        stats["embedded_images_blank"] = 1
    if reason == "IMAGE_DUPLICATE":
        stats["embedded_images_duplicate"] = 1
    return stats


def merge_embedded_image_stats(target, source):
    """Merge embedded image stats into the document stats bucket."""

    for key in [
        "embedded_images_recognized",
        "embedded_images_skipped",
        "embedded_images_duplicate",
        "embedded_images_blank",
        "images_skipped",
        "images_blank",
        "images_duplicate",
        "images_compressed",
        "images_recognized",
    ]:
        target[key] = int(target.get(key) or 0) + int(source.get(key) or 0)


def embedded_images_markdown(items):
    """Return markdown for recognized embedded images."""

    if not items:
        return ""
    parts = ["## Embedded Images"]
    for index, item in enumerate(items, start=1):
        parts.extend(
            [
                "",
                f"### Embedded Image {index}",
                "",
                f"Source: {item['source']}",
                "",
                item["description"].strip(),
            ]
        )
    return "\n".join(parts).strip()


def source_bytes_sha256(data):
    """Return a SHA-256 digest for bytes."""

    import hashlib

    return hashlib.sha256(data).hexdigest()


def convert_pdf_images(path, context):
    """Recognize PDF embedded images and optionally scanned pages."""

    stats = empty_pdf_stats()
    cost = empty_cost_stats()
    descriptions = []
    conversion = context.get("conversion") or {}
    if not vision_configured(context):
        return {"markdown": "", "stats": stats, "cost": cost}
    try:
        import fitz
    except Exception:
        stats["pdf_images_skipped"] += 1
        return {"markdown": "", "stats": stats, "cost": cost}

    options = pdf_options(conversion)
    seen = set()
    assets_dir = sidecar_path(path) / "assets"
    try:
        with fitz.open(path) as document:
            stats["pdf_pages"] = int(document.page_count or 0)
            page_count = min(stats["pdf_pages"], options["max_pages"])
            for page_index in range(page_count):
                page = document.load_page(page_index)
                stats["pdf_pages_processed"] += 1
                page_has_text = pdf_page_has_text(
                    page,
                    options["min_text_chars"],
                )
                if page_has_text:
                    stats["pdf_pages_with_text"] += 1
                else:
                    stats["pdf_scanned_pages"] += 1
                    if options["render_scanned_pages"]:
                        if stats["pdf_images_total"] >= options["max_images"]:
                            result = pdf_skipped_result(
                                page_index,
                                "CONVERSION_MAX_IMAGES_EXCEEDED",
                                meta={"kind": "rendered_page"},
                            )
                        else:
                            result = convert_pdf_rendered_page(
                                path,
                                assets_dir,
                                page,
                                page_index,
                                options,
                                context,
                            )
                        merge_pdf_item_stats(stats, result["stats"])
                        merge_cost_stats(cost, result["cost"])
                        if result["description"]:
                            descriptions.append(result)
                should_extract_images = (
                    options["extract_images"]
                    and (
                        not page_has_text
                        or options["extract_images_on_text_pages"]
                    )
                )
                if should_extract_images:
                    results = convert_pdf_page_images(
                        path,
                        assets_dir,
                        document,
                        page,
                        page_index,
                        options,
                        context,
                        seen,
                        max(
                            0,
                            options["max_images"] - stats["pdf_images_total"],
                        ),
                    )
                    for result in results:
                        merge_pdf_item_stats(stats, result["stats"])
                        merge_cost_stats(cost, result["cost"])
                        if result["description"]:
                            descriptions.append(result)
    except Exception:
        stats["pdf_images_skipped"] += 1

    return {
        "markdown": pdf_images_markdown(descriptions),
        "stats": stats,
        "cost": cost,
    }


def empty_pdf_stats():
    """Return empty PDF image recognition stats."""

    return {
        "pdf_pages": 0,
        "pdf_pages_processed": 0,
        "pdf_pages_with_text": 0,
        "pdf_scanned_pages": 0,
        "pdf_images_total": 0,
        "pdf_images_recognized": 0,
        "pdf_images_skipped": 0,
        "pdf_rendered_pages": 0,
    }


def pdf_options(conversion):
    """Return PDF image conversion options."""

    return {
        "extract_images": conversion.get("pdf_extract_images") is not False,
        "extract_images_on_text_pages": bool(
            conversion.get("pdf_extract_images_on_text_pages")
        ),
        "max_images": int(conversion.get("max_images") or DEFAULT_MAX_IMAGES),
        "max_images_per_page": int(
            conversion.get("pdf_max_images_per_page")
            or DEFAULT_PDF_MAX_IMAGES_PER_PAGE
        ),
        "max_pages": int(
            conversion.get("pdf_max_pages") or DEFAULT_PDF_MAX_PAGES
        ),
        "min_text_chars": int(
            conversion.get("pdf_min_text_chars")
            or DEFAULT_PDF_MIN_TEXT_CHARS
        ),
        "min_image_area_ratio": float(
            conversion.get("pdf_min_image_area_ratio")
            or DEFAULT_PDF_MIN_IMAGE_AREA_RATIO
        ),
        "render_dpi": int(
            conversion.get("pdf_render_dpi") or DEFAULT_PDF_RENDER_DPI
        ),
        "render_scanned_pages": bool(
            conversion.get("pdf_render_scanned_pages")
        ),
    }


def pdf_page_has_text(page, min_text_chars):
    """Return whether a PDF page has enough extractable text."""

    try:
        text = page.get_text("text") or ""
    except Exception:
        return False
    return len(text.strip()) >= min_text_chars


def convert_pdf_page_images(
    source_path,
    assets_dir,
    document,
    page,
    page_index,
    options,
    context,
    seen,
    remaining_images,
):
    """Recognize embedded images from one PDF page."""

    results = []
    images = page.get_images(full=True) or []
    for image_index, image in enumerate(images, start=1):
        if remaining_images <= 0:
            results.append(
                pdf_skipped_result(
                    page_index,
                    "CONVERSION_MAX_IMAGES_EXCEEDED",
                )
            )
            continue
        if image_index > options["max_images_per_page"]:
            results.append(
                pdf_skipped_result(page_index, "PDF_PAGE_IMAGE_LIMIT")
            )
            continue
        if not pdf_image_has_enough_page_area(page, image, options):
            results.append(
                pdf_skipped_result(page_index, "PDF_IMAGE_TOO_SMALL")
            )
            continue
        xref = int(image[0])
        try:
            extracted = document.extract_image(xref)
        except Exception:
            results.append(pdf_skipped_result(page_index, "PDF_IMAGE_INVALID"))
            continue
        raw = extracted.get("image") or b""
        if not raw:
            results.append(pdf_skipped_result(page_index, "PDF_IMAGE_EMPTY"))
            continue
        digest = source_bytes_sha256(raw)
        if digest in seen:
            results.append(pdf_skipped_result(page_index, "IMAGE_DUPLICATE"))
            continue
        seen.add(digest)
        suffix = f".{extracted.get('ext') or 'png'}"
        name = f"pdf_p{page_index + 1}_x{xref}{suffix}"
        results.append(
            convert_pdf_image_bytes(
                source_path,
                assets_dir,
                name,
                raw,
                context,
                {
                    "page": page_index + 1,
                    "kind": "embedded_image",
                    "source": f"page={page_index + 1}, xref={xref}",
                },
            )
        )
        remaining_images -= 1
    return results


def pdf_image_has_enough_page_area(page, image, options):
    """Return whether a PDF image is large enough on the page."""

    page_area = float(page.rect.width * page.rect.height)
    if page_area <= 0:
        return False
    xref = int(image[0])
    try:
        rects = page.get_image_rects(xref)
    except Exception:
        return True
    if not rects:
        return True
    max_area = max(float(rect.width * rect.height) for rect in rects)
    return max_area / page_area >= options["min_image_area_ratio"]


def convert_pdf_rendered_page(
    source_path,
    assets_dir,
    page,
    page_index,
    options,
    context,
):
    """Render and recognize one scanned PDF page."""

    import fitz

    matrix = fitz.Matrix(
        options["render_dpi"] / 72,
        options["render_dpi"] / 72,
    )
    pixmap = page.get_pixmap(matrix=matrix, alpha=False)
    raw = pixmap.tobytes("png")
    name = f"pdf_p{page_index + 1}_rendered.png"
    return convert_pdf_image_bytes(
        source_path,
        assets_dir,
        name,
        raw,
        context,
        {
            "page": page_index + 1,
            "kind": "rendered_page",
            "source": (
                f"page={page_index + 1}, "
                f"render_dpi={options['render_dpi']}"
            ),
            "rendered_page": True,
        },
    )


def convert_pdf_image_bytes(source_path, assets_dir, name, raw, context, meta):
    """Recognize one PDF image-like asset."""

    assets_dir.mkdir(parents=True, exist_ok=True)
    suffix = Path(name).suffix.lower() or ".png"
    digest = source_bytes_sha256(raw)
    asset_path = assets_dir / f"pdf_{digest[:12]}{suffix}"
    asset_path.write_bytes(raw)
    prepared = prepare_image_for_model(asset_path, context)
    if prepared.get("skipped"):
        return pdf_skipped_result(
            int(meta.get("page") or 0) - 1,
            prepared.get("reason") or "IMAGE_SKIPPED",
            prepared.get("stats") or {},
            meta,
        )

    content, usage = describe_image_bytes(
        prepared["bytes"],
        prepared["mime_type"],
        context,
    )
    stats = {
        **(prepared.get("stats") or {}),
        "images_recognized": 1 if content else 0,
        "pdf_images_recognized": 1 if content else 0,
        "pdf_images_total": 1,
    }
    if meta.get("rendered_page"):
        stats["pdf_rendered_pages"] = 1
    return {
        "source": meta["source"],
        "asset": asset_path.name,
        "page": meta.get("page") or 0,
        "kind": meta.get("kind") or "embedded_image",
        "description": content,
        "stats": stats,
        "cost": model_cost_stats(usage, content),
    }


def pdf_skipped_result(page_index, reason, stats=None, meta=None):
    """Return a skipped PDF image result."""

    stats = dict(stats or {})
    stats["pdf_images_total"] = int(stats.get("pdf_images_total") or 0) + 1
    stats["pdf_images_skipped"] = max(
        1,
        int(stats.get("pdf_images_skipped") or 0),
    )
    stats["images_skipped"] = max(1, int(stats.get("images_skipped") or 0))
    return {
        "source": (meta or {}).get("source") or f"page={page_index + 1}",
        "asset": "",
        "page": page_index + 1,
        "kind": (meta or {}).get("kind") or "embedded_image",
        "description": "",
        "reason": reason,
        "stats": stats,
        "cost": empty_cost_stats(),
    }


def merge_pdf_item_stats(target, source):
    """Merge one PDF image result into aggregate document stats."""

    for key in [
        "pdf_images_total",
        "pdf_images_recognized",
        "pdf_images_skipped",
        "pdf_rendered_pages",
        "images_recognized",
        "images_skipped",
        "images_blank",
        "images_duplicate",
        "images_compressed",
    ]:
        target[key] = int(target.get(key) or 0) + int(source.get(key) or 0)


def pdf_images_markdown(items):
    """Return markdown for recognized PDF images."""

    if not items:
        return ""
    parts = ["## PDF Images"]
    for index, item in enumerate(items, start=1):
        title = "Rendered Page" if item["kind"] == "rendered_page" else "Image"
        parts.extend(
            [
                "",
                f"### Page {item['page']} · {title} {index}",
                "",
                f"Source: {item['source']}",
                "",
                item["description"].strip(),
            ]
        )
    return "\n".join(parts).strip()


def image_skip_stats(reason):
    """Return stats for one skipped image."""

    stats = {"images_skipped": 1}
    if reason == "IMAGE_BLANK":
        stats["images_blank"] = 1
    elif reason == "IMAGE_DUPLICATE":
        stats["images_duplicate"] = 1
    return stats


def prepare_image_for_model(path, context):
    """Return optimized image bytes for model recognition."""

    conversion = context.get("conversion") or {}
    min_bytes = int(
        conversion.get("min_image_bytes") or DEFAULT_MIN_IMAGE_BYTES
    )
    if path.stat().st_size < min_bytes:
        return {
            "skipped": True,
            "reason": "IMAGE_TOO_SMALL",
            "stats": image_skip_stats("IMAGE_TOO_SMALL"),
        }

    try:
        from PIL import Image, ImageStat, UnidentifiedImageError
    except Exception:
        return {
            "bytes": path.read_bytes(),
            "mime_type": mimetypes.guess_type(str(path))[0] or "image/png",
            "stats": {"images_preprocess_unavailable": 1},
        }

    try:
        image = Image.open(path)
        image.load()
    except (OSError, UnidentifiedImageError):
        return {
            "skipped": True,
            "reason": "IMAGE_INVALID",
            "stats": image_skip_stats("IMAGE_INVALID"),
        }

    min_dimension = int(
        conversion.get("min_image_dimension") or DEFAULT_MIN_IMAGE_DIMENSION
    )
    width, height = image.size
    if width < min_dimension or height < min_dimension:
        return {
            "skipped": True,
            "reason": "IMAGE_TOO_SMALL",
            "stats": image_skip_stats("IMAGE_TOO_SMALL"),
        }
    if image_is_blank(image, ImageStat):
        return {
            "skipped": True,
            "reason": "IMAGE_BLANK",
            "stats": image_skip_stats("IMAGE_BLANK"),
        }

    return optimize_image_for_model(image, context, path)


def image_is_blank(image, image_stat):
    """Return whether an image is effectively blank."""

    if "A" in image.getbands():
        alpha = image.getchannel("A")
        if alpha.getextrema() == (0, 0):
            return True
    grayscale = image.convert("L")
    stat = image_stat.Stat(grayscale)
    return float(stat.var[0] or 0) < IMAGE_BLANK_VARIANCE_THRESHOLD


def optimize_image_for_model(image, context, path):
    """Downscale and re-encode one image before model upload."""

    conversion = context.get("conversion") or {}
    max_dimension = int(
        conversion.get("image_max_dimension") or DEFAULT_IMAGE_MAX_DIMENSION
    )
    quality = int(
        conversion.get("image_jpeg_quality") or DEFAULT_IMAGE_JPEG_QUALITY
    )
    original_width, original_height = image.size
    original_data = path.read_bytes()
    image = image.copy()
    resized = False
    if max(original_width, original_height) > max_dimension:
        image.thumbnail((max_dimension, max_dimension))
        resized = True

    has_alpha = "A" in image.getbands()
    buffer = BytesIO()
    stats = {
        "image_original_width": original_width,
        "image_original_height": original_height,
        "image_width": image.size[0],
        "image_height": image.size[1],
    }
    if has_alpha:
        image.save(buffer, format="PNG", optimize=True)
        mime_type = "image/png"
    else:
        image.convert("RGB").save(
            buffer,
            format="JPEG",
            optimize=True,
            quality=max(1, min(95, quality)),
        )
        mime_type = "image/jpeg"

    data = buffer.getvalue()
    if not resized and len(data) >= len(original_data):
        data = original_data
        mime_type = mimetypes.guess_type(str(path))[0] or "image/png"
    elif len(data) < len(original_data) or resized:
        stats["images_compressed"] = 1
    stats["image_upload_bytes"] = len(data)
    return {"bytes": data, "mime_type": mime_type, "stats": stats}


def describe_image_file(path, context):
    """Describe an image file through the LensNode gateway."""

    prepared = prepare_image_for_model(path, context)
    if prepared.get("skipped"):
        raise RuntimeError(prepared.get("reason") or "IMAGE_SKIPPED")
    return describe_image_bytes(
        prepared["bytes"],
        prepared["mime_type"],
        context,
    )


def describe_image_bytes(image_bytes, mime_type, context):
    """Describe image bytes through the LensNode gateway."""

    conversion = context.get("conversion") or {}
    model_ref = conversion.get("vision_model_ref") or context.get(
        "vision_model_ref"
    )
    gateway_url = context.get("ai_gateway_url")
    token = context.get("lensnode_token")
    if not model_ref or not gateway_url or not token:
        raise RuntimeError("VISION_MODEL_NOT_CONFIGURED")
    from .gateway_model import describe_image_result

    prompt = image_prompt(context)
    result = describe_image_result(
        image_bytes,
        prompt,
        mime_type,
        model_ref=model_ref,
        ai_gateway_url=gateway_url,
        token=token,
    )
    return result.get("content") or "", result.get("usage") or {}


def vision_configured(context):
    """Return whether image recognition has enough gateway settings."""

    conversion = context.get("conversion") or {}
    return bool(
        (conversion.get("vision_model_ref") or context.get("vision_model_ref"))
        and context.get("ai_gateway_url")
        and context.get("lensnode_token")
    )


def document_image_context(context, path, text):
    """Return a context enriched with document language and text snippet."""

    snippet = compact_text(text, 2000)
    return {
        **context,
        "document_title": Path(path).name,
        "document_context": snippet,
        "document_language": detect_text_language(snippet, Path(path).name),
    }


def standalone_image_context(context, path):
    """Return a context for standalone image files."""

    path_text = str(path)
    return {
        **context,
        "document_title": Path(path).name,
        "document_context": path_text,
        "document_language": detect_text_language("", path_text),
    }


def compact_text(text, limit):
    """Return compact text limited to the given number of characters."""

    value = " ".join(str(text or "").split())
    return value[:limit]


def detect_text_language(text="", fallback=""):
    """Return zh or en based on visible text."""

    sample = f"{text or ''} {fallback or ''}"
    zh_count = 0
    latin_count = 0
    for char in sample:
        if "\u4e00" <= char <= "\u9fff":
            zh_count += 1
        elif ("a" <= char.lower() <= "z"):
            latin_count += 1
    if zh_count >= 2 and zh_count >= latin_count * 0.05:
        return "zh"
    if latin_count > 0:
        return "en"
    return "zh"


def image_prompt(context=None):
    """Return the image description prompt."""

    context = context or {}
    language = context.get("document_language") or "zh"
    title = context.get("document_title") or ""
    snippet = context.get("document_context") or ""
    if language == "en":
        return "\n".join(
            [
                "Describe this image in English for knowledge-base retrieval.",
                "Extract visible text, chart/table data, key objects, and "
                "business meaning. Be specific and avoid generic wording.",
                "If the image is only a logo, icon, watermark, decoration, "
                "or other low-value visual, say so briefly.",
                f"Document title: {title}",
                f"Document context: {snippet}",
            ]
        )
    return "\n".join(
        [
            "请使用简体中文为知识库检索生成图片描述。",
            "请提取图片中的文字、图表/表格数据、关键对象和业务含义，"
            "避免泛泛描述。",
            "如果图片只是 logo、图标、水印、装饰图等低价值视觉元素，"
            "请简短说明即可。",
            f"文档标题：{title}",
            f"文档上下文：{snippet}",
        ]
    )


def content_markdown(target, path, context, text):
    """Return searchable markdown with a compact frontmatter."""

    source_path = path.resolve().relative_to(Path(target).resolve()).as_posix()
    source_name = path.name
    return (
        "---\n"
        f"datasource_uuid: {context.get('datasource_uuid') or ''}\n"
        f"source_type: {context.get('source_type') or ''}\n"
        f"source_path: {source_path}\n"
        f"source_name: {source_name}\n"
        f"converter: {converter_version(path).get('name')}\n"
        "---\n\n"
        f"# {source_name}\n\n"
        f"{text.strip()}\n"
    )


def write_success_meta(
    target,
    path,
    item,
    context,
    digest,
    fingerprint,
    chars,
    images_recognized,
    stats=None,
):
    """Write successful conversion metadata."""

    stats = dict(stats or {})
    stats["chars"] = chars
    stats.setdefault("images_total", images_recognized)
    stats.setdefault("images_recognized", images_recognized)
    write_meta(
        target,
        path,
        item,
        context,
        {
            "status": "success",
            "error": "",
            "fingerprint": fingerprint,
            "stats": stats,
            "source_sha256": digest,
        },
    )


def write_failed_meta(target, path, item, context, error):
    """Write failed conversion metadata."""

    write_meta(
        target,
        path,
        item,
        context,
        {
            "status": "failed",
            "error": error,
            "fingerprint": "",
            "stats": {},
            "source_sha256": "",
        },
    )


def write_skipped_meta(target, path, item, context, reason):
    """Write skipped conversion metadata."""

    write_meta(
        target,
        path,
        item,
        context,
        {
            "status": "skipped",
            "error": reason,
            "fingerprint": "",
            "stats": {},
            "source_sha256": "",
        },
    )


def write_meta(target, path, item, context, conversion):
    """Write sidecar meta.json."""

    sidecar_path(path).mkdir(parents=True, exist_ok=True)
    source_path = path.resolve().relative_to(Path(target).resolve()).as_posix()
    payload = {
        "schema_version": 1,
        "datasource": {
            "uuid": context.get("datasource_uuid") or "",
            "name": context.get("name") or "",
            "source_type": context.get("source_type") or "",
        },
        "source": {
            "path": source_path,
            "name": path.name,
            "extension": path.suffix.lstrip(".").lower(),
            "size": path.stat().st_size if path.exists() else 0,
            "sha256": conversion.pop("source_sha256", ""),
            "mtime": time.strftime(
                "%Y-%m-%dT%H:%M:%SZ",
                time.gmtime(path.stat().st_mtime if path.exists() else 0),
            ),
            "remote": item.get("remote") or {},
        },
        "conversion": {
            "status": conversion.get("status"),
            "error": conversion.get("error", ""),
            "fingerprint": conversion.get("fingerprint", ""),
            "generated_at": time.strftime(
                "%Y-%m-%dT%H:%M:%SZ",
                time.gmtime(),
            ),
            "options": context.get("conversion") or {},
            "tool": converter_version(path),
            "model": {
                "ref": (context.get("conversion") or {}).get(
                    "vision_model_ref",
                    "",
                ),
                "prompt_version": PROMPT_VERSION,
            },
            "stats": conversion.get("stats") or {},
        },
        "retrieval": {
            "content_path": (
                sidecar_path(path)
                / "content.md"
            ).resolve().relative_to(Path(target).resolve()).as_posix(),
            "source_path": source_path,
            "title": safe_filename(path.name),
            "mime_type": mimetypes.guess_type(str(path))[0] or "",
        },
    }
    (sidecar_path(path) / "meta.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def read_json(path):
    """Read a JSON file or return empty dict."""

    try:
        return json.loads(Path(path).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def is_document_path(path):
    """Return whether a path is a supported document."""

    return Path(path).suffix.lower() in DOCUMENT_EXTENSIONS


def is_image_path(path):
    """Return whether a path is a supported image."""

    return Path(path).suffix.lower() in IMAGE_EXTENSIONS


def emit_conversion(
    emit,
    step,
    status,
    message,
    summary,
    total=0,
    current=0,
    current_file="",
    current_status="",
    current_reason="",
    current_stats=None,
):
    """Emit one conversion progress event."""

    if emit is None:
        return
    percent = 100 if total <= 0 else int(current * 100 / total)
    event_summary = dict(summary or {})
    if step != "conversion_manifest":
        event_summary.pop("items", None)
        event_summary.pop("items_truncated", None)
    emit(
        {
            "step": step,
            "status": status,
            "message": message,
            "category": "conversion",
            "summary": event_summary,
            "progress_total": total,
            "progress_current": current,
            "progress_percent": max(0, min(100, percent)),
            "current_file": current_file,
            "current_status": current_status,
            "current_reason": current_reason,
            "current_stats": current_stats or {},
        }
    )
